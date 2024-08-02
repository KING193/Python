# Files Excel
import os
import openpyxl

# pip install openpyxl

os.chdir("C:\\Users\\DELL\\Desktop\\Programr\\Python\\iS7ee7\\EXCEL")

workbook = openpyxl.load_workbook("ExcelSecret.xlsx")
print(workbook.sheetnames)# => ['Feuil1', 'Feuil2']

Sheet1 = workbook["Feuil1"]
print(Sheet1["A1"].value)# => Names
print(Sheet1["A2"].value)# => osama
print(Sheet1["A3"].value)# => khaled

Sheet1["A5"].value = "ibrahim"
Sheet1["B5"].value = "33"
workbook.save("Excel-secret.xlsx")

print(Sheet1.cell(row=1, column=1))# => <Cell 'Feuil1'.A1>
print(Sheet1.cell(row=1, column=1).value)# => Names

for i in range(1, 6):
	print(Sheet1.cell(row=i, column=1).value)# => Names
											  # => osama
											  # => khaled
											  # => mohammed
											  # => ibrahim

for i in range(1, 6):
	print(Sheet1.cell(row=i, column=2).value)# => Ages
											  # => 18
											  # => 33
											  # => 32
											  # => 33

for i in range(1, 6):
	print(Sheet1.cell(row=i, column=1).value, Sheet1.cell(row=i, column=2).value)# => Names Ages
											  									 # => osama 18
											  									 # => khaled 33
											  									 # => mohammed 32
											  									 # => ibrahim 33