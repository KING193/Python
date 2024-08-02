# Files Excel
import os
import openpyxl

os.chdir("C:\\Users\\DELL\\Desktop\\Programr\\Python\\iS7ee7\\EXCEL")

workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.cell(row=1, column=1).value = "Names"
sheet1.cell(row=1, column=2).value = "Ages"
workbook.save("ExcelSecret-2.xlsx")

# python openpyxl

#   workbook
#   sheets or feuil
#   cell