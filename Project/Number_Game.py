import random
import termcolor

import os
import openpyxl

first_name = input("Enter first name: ")
last_name = input("Enter last name: ")

start = "\nWelcome " + str(termcolor.colored(first_name + " " + last_name, "blue")) + " in NUMBER GAME!!\n"
print(start)

point = 0
round = 1

win = termcolor.colored("You Win!", "green")
lost = termcolor.colored("You Lost", "red")

while round <= 10:
    number = random.randint(1, 10)

    y_number = "random number is: " + str(termcolor.colored(number, "yellow"))

    try:
        x_number = int(input("\nEnter the number: "))
        
        if round == 1:
        
            if number == x_number:
                
                print(win)
                round += 1
                point = 100

            else:
                
                print(lost)
                round += 1
                print(y_number)

        elif round == 10:
            
            if number == x_number:
                
                print(win)
                round += 1
                point += 26

            else:
                
                print(lost)
                round += 1
                print(y_number)

        else:

            if number == x_number:
                
                print(win)
                round += 1
                point += 50

            else:
                
                print(lost)
                round += 1
                print(y_number)

    except:
        print(termcolor.colored("Error!!!\nPlease Enter The Numbers Only", "black"))

print("\nTotal point is " + str(termcolor.colored(point, "blue")))

os.chdir("C:\\Users\\DELL\\Desktop\\Programr\\Python\\Project")
workbook = openpyxl.Workbook()

sheet = workbook.active
sheet["A1"].value = "NAMES"
sheet["B1"].value = "SCORE"

sheet.cell(row=2, column=1).value = first_name + " " + last_name
sheet.cell(row=2, column=2).value = point

workbook.save("Player List.xlsx")